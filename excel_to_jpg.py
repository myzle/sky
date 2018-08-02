# coding:utf-8

from openpyxl import load_workbook
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
from math import floor, ceil
import httplib2
from datetime import *
import os
from random import sample
from time import sleep


# excelname为绝对路径传入，columns为想要打印的列，从文件中读数据，打印成图片
# excelname = '/Users/hecom/Desktop/红圈通巡检结果/红圈通巡检报告201705251023.xlsx',columns=[1,4,5]
def excel_to_jpg(excelname, columns):
    # 读excel之前的准备工作
    if excelname == '' or not os.path.exists(excelname):
        result = 'excel name error'
        return result

    try:
        excel_document = load_workbook(excelname)
    except FileNotFoundError as e:
        result = 'open excel error'
        return result

    try:
        sheet_name = excel_document.get_sheet_names()[0]
        sheet = excel_document.get_sheet_by_name(sheet_name)
    except KeyError as e:
        result = 'open sheet error'
        return result

    # 设置图片基本信息，单元格宽度和高度，划线宽度
    # cell_height = 150
    # cell_width = 900
    cell_height = 75
    cell_width = 450
    line_width = 5

    # 基本字体
    char_style = '/Users/myzle/PycharmProjects/sky/Kaiti.ttc'

    # 字体大小
    # char_size = 90
    # title_char_size = 120

    char_size = 45
    title_char_size = 60

    # 颜色
    table_color = (255, 255, 255)
    line_color = (225, 225, 225)
    cell_color = (225, 225, 225)
    title_cell_color = (200, 200, 200)
    char_color = (0, 0, 0)
    error_char_color = (220, 20, 60)

    title_font = ImageFont.truetype(font=char_style, size=title_char_size, index=0, encoding='')
    char_font = ImageFont.truetype(font=char_style, size=char_size, index=0, encoding='')

    # 获取图片大小
    rows_number = sheet.max_row
    columns_number = len(columns)

    table_height = cell_height * rows_number
    table_width = cell_width * columns_number

    # 设置缩放比例

    # 画图片
    im = Image.new('RGB', (table_width, table_height), table_color)
    draw = ImageDraw.Draw(im)

    for y in range(0, rows_number + 1):
        draw.line(((0, y * cell_height), (table_width, y * cell_height)), fill=line_color, width=line_width)

    for x in range(0, columns_number + 1):
        draw.line(((x * cell_width, 0), (x * cell_width, table_height)), fill=line_color, width=line_width)

    draw.rectangle(((0, 0), (table_width, cell_height)), fill=title_cell_color)

    # 写标题
    title = sheet['A1'].value + '(' + os.path.splitext(excelname)[0][-12:] + ')'

    x = floor(cell_width * columns_number / 4)
    draw.text((x, 0), str(title), fill=char_color, font=title_font)

    # 读数据
    datas = []
    if rows_number > 1:
        for i in range(2, rows_number + 1):
            for j in columns:
                data = sheet.cell(row=i, column=j).value
                datas.append(data)

    # print(datas)

    # 写数据
    if rows_number > 1:
        for j in range(1, rows_number):
            for i in range(0, columns_number):
                write_char = datas.pop(0)
                if (write_char == None):
                    pass
                else:
                    write_char = str(write_char)
                    if (write_char.find('fail') == -1 and write_char.find('rror') == -1 and write_char.find(
                            'timeout') == -1 and write_char.find('Exception') == -1):
                        draw.text((cell_width * i, cell_height * j), str(write_char), fill=char_color, font=char_font)
                    else:
                        draw.text((cell_width * i, cell_height * j), str(write_char), fill=error_char_color,
                                  font=char_font)

    # create the result folder
    save_path = os.path.split(excelname)[0] + '/巡检结果图片/'
    # print(save_path)

    if (os.path.exists(save_path)):
        pass
    else:
        os.mkdir(save_path)

    src = 'ABCEDFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'

    jpg_name = os.path.split(os.path.splitext(excelname)[0])[1][-12:] + '.jpg'

    # print(jpg_name)

    pic_full_path = save_path + jpg_name

    im.save(pic_full_path)
    print(pic_full_path)

    return (pic_full_path, jpg_name)


# 参数为excel_to_jpg返回的结果
def get_jpg_url_name_and_send_to_group(receiveGroupId, jpg_full_path_name, jpg_name):

    print('发送图片')
    # 传送图片的ip地址改一下
    root_server_address = 'hecom@123.57.30.193:/usr/share/nginx/html/jietu/website'

    results = os.popen('scp {0} {1}'.format(jpg_full_path_name, root_server_address))

    # 等待图片传送完毕，以后运行可能因为网络慢，导致图片上传速度慢，这里出现问题
    sleep(2)

    # 调用开发接口把图片传到群里去
    jpg_url = 'http://123.57.30.193/website/' + jpg_name

    httplib2.debuglevel = 4

    h = httplib2.Http()

    # url_data = 'http://220.181.105.91:9888/mobile-0.0.1-SNAPSHOT/rcm/common/user/sendGroupImage.do?userStr={"receiveGroupId":"246889940502187140","sendUid":"rcmuser7002","url":"http://sucai.qqjay.com/qqjayxiaowo/201210/26/1.jpg","key":"xiaojiejie654321"}'

    # 测试红圈通的url，不要带汉字，必须是阿里云等的公网，不能是公司自己内部搭的服务器。否则无workout: 因为内网和外网的原因
    # receiveGroupId：接收群的id，sendUid：发送人的id，需要开发在后台数据库中查询才能得到。
    send_jpg_url = jpg_url
    # url_data = 'http://220.181.105.91:9888/mobile-0.0.1-SNAPSHOT/rcm/common/user/sendGroupImage.do?userStr={"receiveGroupId":"17075459391489","sendUid":"rcmuser1826082","url":"' + send_jpg_url + '","key":"xiaojiejie654321"}'
    url_data = 'https://mm.hecom.cn/mobile-0.0.1-SNAPSHOT/rcm/common/user/sendGroupImage.do?userStr={"receiveGroupId":"' + receiveGroupId +'","sendUid":"rcmuser10062739","url":"' + send_jpg_url + '","key":"xiaojiejie654321"}'

    r, c = h.request(uri=url_data)

    if (r.status == 429 or r.status == 503):
        print('被限流')
        return 'error'

        # r,c = h.request(uri = url, method = 'POST', body = data)
        # r,c = h.request('https://www.baidu.com')
        # print(r.status)
        # print(c)

    return 'success'


def generate_jpg_from_excel_and_send_to_group(receiveGroupId, excelname, columns):
    jpg_full_path_name, jpg_name = excel_to_jpg(excelname, columns)
    get_jpg_url_name_and_send_to_group(receiveGroupId, jpg_full_path_name, jpg_name)


if __name__ == "__main__":
    # result = excel_to_jpg('/Users/hecom/Desktop/红圈通巡检结果/红圈通巡检报告201705141023.xlsx',columns=[1,4,5])
    generate_jpg_from_excel_and_send_to_group(excelname='/Users/hecom/Desktop/红圈通巡检结果/红圈通巡检报告201705251023.xlsx',
                                              columns=[1, 4, 5])
    # print(result)

